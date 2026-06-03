"""批量生成 Excel 文件，然后将所有数据导入 SQLite 数据库。"""

from pathlib import Path
import random
import sqlite3
import string
import time

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
EXCEL_DIR = BASE_DIR / "exp4_generated_excel"
DB_PATH = BASE_DIR / "exp4_lab4_data.db"
FILE_COUNT = 50
HEADERS = ["col1", "col2", "col3", "col4", "col5"]


def random_text(length=8):
    """生成指定长度的随机字母数字字符串。"""
    chars = string.ascii_letters + string.digits
    return "".join(random.choice(chars) for _ in range(length))


def generate_excel_files(output_dir=EXCEL_DIR, file_count=FILE_COUNT):
    """批量生成含随机数据的 Excel 文件。"""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    row_counts = random.sample(range(20, 200), file_count)
    paths = []
    for index, row_count in enumerate(row_counts, 1):
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(HEADERS)
        for _ in range(row_count):
            ws.append([random_text() for _ in HEADERS])

        path = output_dir / f"data_{index:02d}.xlsx"
        wb.save(path)
        paths.append(path)
    return paths


def create_table(db_path=DB_PATH):
    """在 SQLite 中创建目标表（若已存在则先删除）。"""
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute("DROP TABLE IF EXISTS excel_records")
        conn.execute(
            """
            CREATE TABLE excel_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file TEXT NOT NULL,
                row_no INTEGER NOT NULL,
                col1 TEXT,
                col2 TEXT,
                col3 TEXT,
                col4 TEXT,
                col5 TEXT
            )
            """
        )


def iter_excel_rows(excel_dir):
    """遍历目录下所有 Excel 文件，逐行生成 (文件名, 行号, 值列表)。"""
    excel_dir = Path(excel_dir)
    for path in sorted(excel_dir.glob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        try:
            for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                values = list(row[:5])
                while len(values) < 5:
                    values.append(None)
                yield path.name, row_no, values
        finally:
            wb.close()


def import_excel_files(excel_dir=EXCEL_DIR, db_path=DB_PATH):
    """将所有 Excel 数据批量插入 SQLite，返回插入行数。"""
    rows = (
        (source_file, row_no, *values)
        for source_file, row_no, values in iter_excel_rows(excel_dir)
    )
    with sqlite3.connect(db_path) as conn:
        cursor = conn.executemany(
            """
            INSERT INTO excel_records
                (source_file, row_no, col1, col2, col3, col4, col5)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()
        return cursor.rowcount


def main():
    generate_excel_files(EXCEL_DIR, FILE_COUNT)
    create_table(DB_PATH)

    start = time.perf_counter()
    total = import_excel_files(EXCEL_DIR, DB_PATH)
    elapsed = time.perf_counter() - start
    speed = total / elapsed if elapsed else 0

    print(f"已生成 {FILE_COUNT} 个Excel文件：{EXCEL_DIR}")
    print(f"已导入 {total} 条记录到数据库：{DB_PATH}")
    print(f"导入耗时：{elapsed:.4f} 秒")
    print(f"平均导入速度：{speed:.2f} 条/秒")


if __name__ == "__main__":
    main()
