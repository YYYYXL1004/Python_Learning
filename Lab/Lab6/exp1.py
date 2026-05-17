import os
import random

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, GradientFill


# 保存生成的Excel文件的文件夹
SAVE_DIR = os.path.join(os.path.dirname(__file__), "exp1_files")
FILE_COUNT = 3
ROW_COUNT = 10


def make_test_data(file_count):
    """生成指定数量的Excel文件，并写入测试数据。"""
    # 如果保存文件夹不存在，就先创建文件夹
    if not os.path.exists(SAVE_DIR):
        os.mkdir(SAVE_DIR)

    headers = ["学号", "姓名", "语文", "数学", "英语", "总分"]
    names = ["张三", "李四", "王五", "赵六", "钱七", "孙八"]

    for i in range(1, file_count + 1):
        # 创建一个新的工作簿，并设置工作表名称
        wb = Workbook()
        ws = wb.active
        ws.title = "学生成绩"
        ws.append(headers)

        # 随机生成若干行学生成绩数据
        for j in range(1, ROW_COUNT + 1):
            chinese = random.randint(60, 100)
            math = random.randint(60, 100)
            english = random.randint(60, 100)
            ws.append([
                f"2026{i:02d}{j:03d}",
                random.choice(names),
                chinese,
                math,
                english,
                chinese + math + english
            ])

        # 每个工作簿保存为一个单独的Excel文件
        filename = os.path.join(SAVE_DIR, f"test_{i}.xlsx")
        wb.save(filename)


def change_excel_style():
    """批量修改生成的Excel文件格式。"""
    # 设置表头、偶数行、奇数行需要使用的字体和填充样式
    header_font = Font(name="黑体", bold=True, color="FF000000")
    even_font = Font(name="宋体", color="FFFF0000")
    odd_font = Font(name="宋体", color="FF00B0F0")
    even_fill = GradientFill(type="linear", stop=("FFFF0000", "FF0000FF"))

    # 遍历文件夹中所有xlsx文件，逐个打开并修改格式
    for filename in os.listdir(SAVE_DIR):
        if not filename.endswith(".xlsx"):
            continue

        file_path = os.path.join(SAVE_DIR, filename)
        wb = load_workbook(file_path)
        ws = wb.active

        # 第一行是表头，设置为黑体并加粗
        for cell in ws[1]:
            cell.font = header_font

        # 从第二行开始，根据行号奇偶设置不同的字体和背景
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.row % 2 == 0:
                    cell.font = even_font
                    cell.fill = even_fill
                else:
                    cell.font = odd_font

        # 保存修改后的Excel文件
        wb.save(file_path)


def main():
    make_test_data(FILE_COUNT)
    change_excel_style()
    print(f"已经生成并修改{FILE_COUNT}个Excel文件，保存位置：{SAVE_DIR}")


if __name__ == "__main__":
    main()
