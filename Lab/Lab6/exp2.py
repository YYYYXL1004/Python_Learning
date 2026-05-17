import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


# 保存本实验所需Excel文件的文件夹
SAVE_DIR = os.path.join(os.path.dirname(__file__), "exp2_files")
RESULT_FILE = os.path.join(SAVE_DIR, "result.xlsx")
SOURCE_FILES = ["体育.xlsx", "数学.xlsx", "计算机.xlsx"]


def make_excel_file(filename, academy, students, need_merge=True):
    """生成一个学院成绩表。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩表"

    ws.append(["学院", "姓名", "成绩"])

    for name, score in students:
        ws.append([academy, name, score])

    # 为了模拟题目中的情况，不同文件第一列的合并方式可以不同
    if need_merge:
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=len(students) + 1, end_column=1)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    wb.save(os.path.join(SAVE_DIR, filename))


def make_test_files():
    """准备多个具有相同表头结构的Excel文件。"""
    if not os.path.exists(SAVE_DIR):
        os.mkdir(SAVE_DIR)

    data = {
        "体育.xlsx": ("体育", [("a", 80), ("b", 70), ("c", 88), ("g", 79), ("m", 87)], False),
        "数学.xlsx": ("数学", [("a", 88), ("b", 78), ("c", 87), ("d", 68), ("e", 90), ("f", 67), ("g", 80)], True),
        "计算机.xlsx": ("计算机", [("张三", 80), ("李四", 90), ("王五", 70), ("赵六", 85), ("马七", 75)], True)
    }

    for filename, (academy, students, need_merge) in data.items():
        make_excel_file(filename, academy, students, need_merge)


def get_cell_value(ws, row, col):
    """读取单元格内容，如果是合并单元格，就读取左上角单元格的值。"""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    # 循环查找当前单元格属于哪个合并区域
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row,
                           column=merged_range.min_col).value
    else:
        return None


def read_excel_data(file_path):
    """读取一个Excel文件中的表头和数据。"""
    wb = load_workbook(file_path)
    ws = wb.active

    # 使用列表推导式读取表头
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    rows = []

    for row in range(2, ws.max_row + 1):
        # 使用列表推导式读取每一行数据
        line = [get_cell_value(ws, row, col) for col in range(1, ws.max_column + 1)]
        rows.append(line)

    return headers, rows


def merge_excel_files():
    """合并多个Excel文件，并重新合并结果表中第一列相同的学院。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "合并结果"

    # 按指定顺序合并，保证结果和题目示例一致
    file_names = [name for name in SOURCE_FILES if os.path.exists(os.path.join(SAVE_DIR, name))]

    headers = None
    all_rows = []

    for filename in file_names:
        file_path = os.path.join(SAVE_DIR, filename)
        cur_headers, cur_rows = read_excel_data(file_path)

        if headers is None:
            headers = cur_headers
        elif cur_headers != headers:
            print(f"{filename}的表头不一致，不能合并！")
            break

        all_rows.extend(cur_rows)
    else:
        # 只有循环没有被break打断，才会执行这里的合并操作
        ws.append(headers)

        for row in all_rows:
            ws.append(row)

        # 使用生成器推导式统计有效数据行数
        data_count = sum(1 for row in all_rows if row)
        print(f"共合并{len(file_names)}个文件，{data_count}行数据。")

        merge_same_academy(ws)
        set_style(ws)
        wb.save(RESULT_FILE)
        print(f"合并结果已保存到：{RESULT_FILE}")


def merge_same_academy(ws):
    """把结果表中第一列连续相同的学院合并。"""
    start_row = 2

    while start_row <= ws.max_row:
        academy = ws.cell(row=start_row, column=1).value
        end_row = start_row

        while end_row + 1 <= ws.max_row and ws.cell(row=end_row + 1, column=1).value == academy:
            end_row += 1

        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

        start_row = end_row + 1


def set_style(ws):
    """设置结果表的简单格式。"""
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ["A", "B", "C"]:
        ws.column_dimensions[col].width = 12


def main():
    make_test_files()
    merge_excel_files()


if __name__ == "__main__":
    main()
