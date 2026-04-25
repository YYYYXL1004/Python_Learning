import os
import random
import openpyxl

def generate_random_data(filename):
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考试记录"

    # 写入表头
    ws.append(['姓名', '课程', '成绩'])

    names = ['张三', '李四', '王五', '赵六']
    courses = ['Python程序设计', '数据结构与算法', '生成模型基础']

    # 模拟生成100条成绩记录
    for _ in range(100):
        name = random.choice(names)
        course = random.choice(courses)
        score = random.randint(0, 100)
        ws.append([name, course, score])

    wb.save(filename)
    print(f"模拟数据已经保存在 {filename}")


def main(input_filename, output_filename):
    if not os.path.exists(input_filename):
        print(f"找不到文件{input_filename}")
        return
    
    wb_in = openpyxl.load_workbook(input_filename)
    ws_in = wb_in.active

    # 用 Tuple(姓名，课程)作为字典的key
    max_scores = {}

    # iter_rows()按行迭代 min_row=2跳过首行表头， values_only直接返回单元格值而不是对象
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        name, course, score = row

        if not name or not course or score is None:
            continue
        
        key = (name, course)
        max_scores[key] = max(max_scores.get(key, 0), score)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "期末最高成绩"
    ws_out.append(['姓名', '课程', '最高成绩'])

    # 遍历字典，一块写成列表写入
    for (name, course), score in max_scores.items():
        ws_out.append([name, course, score])

    wb_out.save(output_filename)
    print(f"最高分统计完毕，结果保存至: {output_filename}")

if __name__ == "__main__":
    source_file = "all_score.xlsx"
    target_file = "final_score.xlsx"

    generate_random_data(source_file)
    main(source_file, target_file)