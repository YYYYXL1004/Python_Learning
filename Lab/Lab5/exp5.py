import os
import openpyxl
import itertools
from functools import reduce

def generate_mock_excel(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['电影名称', '导演', '演员'])

    # 模拟图片中的前 18 条记录
    mock_data = [
        ["电影1", "导演1", "演员1, 演员2, 演员3, 演员4"],
        ["电影2", "导演2", "演员3, 演员2, 演员4, 演员5"],
        ["电影3", "导演3", "演员1, 演员5, 演员3, 演员6"],
        ["电影4", "导演1", "演员1, 演员4, 演员3, 演员7"],
        ["电影5", "导演2", "演员1, 演员2, 演员3, 演员8"],
        ["电影6", "导演3", "演员5, 演员7, 演员3, 演员9"],
        ["电影7", "导演4", "演员1, 演员4, 演员6, 演员7"],
        ["电影8", "导演1", "演员1, 演员4, 演员3, 演员8"],
        ["电影9", "导演2", "演员5, 演员4, 演员3, 演员9"],
        ["电影10", "导演3", "演员1, 演员4, 演员5, 演员10"],
        ["电影11", "导演1", "演员1, 演员4, 演员3, 演员11"],
        ["电影12", "导演2", "演员7, 演员4, 演员9, 演员12"],
        ["电影13", "导演3", "演员1, 演员7, 演员3, 演员13"],
        ["电影14", "导演4", "演员10, 演员4, 演员9, 演员14"],
        ["电影15", "导演5", "演员1, 演员8, 演员11, 演员15"],
        ["电影16", "导演6", "演员14, 演员4, 演员13, 演员16"],
        ["电影17", "导演7", "演员3, 演员4, 演员9"],
        ["电影18", "导演8", "演员3, 演员4, 演员10"]
    ]

    for row in mock_data:
        ws.append(row)
    wb.save(filename)

def main(input_filename, n):
    if not os.path.exists(input_filename):
        print("找不到文件 {input_filename}")
        return
        
    wb_in = openpyxl.load_workbook(input_filename)
    ws_in = wb_in.active
    
    # 字典结构：{演员名: set(参演电影)}
    actor_movies = {}
    
    for index, row in enumerate(ws_in.iter_rows(values_only=True)):
        if index == 0:  # 跳过首行表头
            continue
            
        movie_name, _, actors_str = row
        if not actors_str:
            continue
            
        # 兼容处理中英文逗号，然后去除两端空格
        actors = [a.strip() for a in actors_str.replace('，', ',').split(',')]
        
        for actor in actors:
            # 获取当前演员的电影集合，如果没有则返回空集合，然后加入当前电影
            movies = actor_movies.get(actor, set())
            movies.add(movie_name)
            actor_movies[actor] = movies

    all_actors = list(actor_movies.keys())
    
    # combinations 生成所有可能的 n 人组合
    combinations = itertools.combinations(all_actors, n)
    
    # 核心逻辑：找出共同参演电影最多的组合
    # 1. [actor_movies[a] for a in combo] 提取这 n 人的电影集合列表
    # 2. reduce(lambda x, y: x & y, ...) 对列表中的集合连续求交集
    # 3. key=lambda combo: len(...) 根据交集的大小选出最大的一组
    best_combo = max(
        combinations,
        key=lambda combo: len(reduce(lambda x, y: x & y, [actor_movies[actor] for actor in combo]))
    )
    
    # 重新算出这组人的共同电影集合，方便打印展示
    common_movies = reduce(lambda x, y: x & y, [actor_movies[actor] for actor in best_combo])
    
    print("=== 统计结果 ===")
    print("关系最好的 {} 个演员是: {}".format(n, ", ".join(best_combo)))
    print("他们共同参演的电影数量为: {}".format(len(common_movies)))
    print("他们共同参演的电影包括: {}".format(common_movies))


if __name__ == "__main__":
    source_file = "电影导演演员.xlsx"
    target_n = 3  # 可以任意修改为 >2 的整数
    
    if not os.path.exists(source_file):
        generate_mock_excel(source_file)
        
    main(source_file, target_n)